#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json
import logging
import hashlib
import openpyxl

from settings import COMMON_URL
from openpyxl.styles import Alignment
from tornado import web, gen, httpclient
from tornado.httputil import url_concat
from settings import APIS, MYKTV_SECRET, SECURET


httpclient.AsyncHTTPClient.configure('tornado.simple_httpclient.SimpleAsyncHTTPClient', max_clients=300)


class APIError(web.HTTPError):
    '''
    自定义API异常
    '''
    def __init__(self, status_code=200, *args, **kwargs):
        super(APIError, self).__init__(status_code, *args, **kwargs)
        self.kwargs = kwargs


def toNone(va):
    # empty to none
    return None if not va else va

def http_request(url, method='GET', **wargs):
    return httpclient.HTTPRequest(url=url, method=method, connect_timeout=10, request_timeout=10, **wargs)

def get_async_client():
    http_client = httpclient.AsyncHTTPClient()
    return http_client

def wxorder_action_to_text(action):
    if action == 0:
        return '扫码支付'
    elif action == 1:
        return '刷卡支付'
    elif action == 2:
        return '公众号支付'

def wxorder_state_to_text(state):
    if state == 0:
        return '订单创建'
    elif state == 1:
        return '订单未支付'
    elif state == 2:
        return '订单已支付'
    elif state == 3:
        return '订单撤销'
    elif state == 4:
        return '订单关闭'
    elif state == 5:
        return '订单退款申请'
    elif state == 6:
        return '订单退款成功'
    elif state == 7:
        return '订单退款失败'

async def fetch(http_client, request):
    r = await http_client.fetch(request)
    logging.info('\treq_url=%s\trequest_time=%s' % (r.effective_url, r.request_time))
    logging.info('\tbody=%s' % (r.body))
    return r

async def async_common_api(path, params={}):
    url = url_concat(COMMON_URL + path, params)
    http_client = get_async_client()
    try:
        request = http_request(url)
        response = await fetch(http_client, request)
        response = json.loads(response.body.decode())
        return response
    except Exception as e:
        logging.error('url=%s, error=%s' % (url, e))
        raise APIError(errcode=10001, errmsg='公共接口请求失败')

def export_xlsx(data, export_filename):
    '''
    data = [
        {'sheetname': 'sheet1', titles: ['title_1', 'title_2', 'title_3'], data: [[1,2,3], [2,3,4], [3,4,5]]},
        {'sheetname': 'sheet2', titles: ['title_1', 'title_1', 'title_1'], data: [[1,2,3], [2,3,4], [3,4,5]]}
    ]
    data是数组，如果长度大于1，则有多少个sheet
    '''
    assert isinstance(data, list)
    wb = openpyxl.Workbook()
    alignment = Alignment(wrap_text=True)
    ws_num = len(data)
    wss = []
    max_len = []   # 是数组的数组: [[], [], []], 记录每一个sheet, 每一列的最大长度, 导出时, 显示更正常
    for i in range(ws_num):
        if i == 0:
            ws = wb.active  # 第一个sheet是这么取的, 如果直接create_sheet, 生成的第一个sheet是空的
            ws.title = data[i].get('sheetname', '')
        else:
            ws = wb.create_sheet(data[i].get('sheetname', ''))
        for idx, title in enumerate(data[i].get('titles', [])):
            col = ord('A') + idx
            ws['%s1' % chr(col)] = title     # 写入标题
        # 初始化每一个sheet的每一列最大宽度为这个列名(汉字)的长度
        max_len.append([len(bytes(str(title), 'GBK')) for title in data[i].get('titles', [])])
        wss.append(ws)
    for idx, ws in enumerate(wss):
        data_lines = data[idx].get('data', [])
        for line, data_line in enumerate(data_lines):
            for col, data_col in enumerate(data_line):
                cur_col = ord('A') + col
                data_col = str(data_col)
                ws['%s%s' % (chr(cur_col), line+2)] = data_col
                ws['%s%s' % (chr(cur_col), line+2)].alignment = alignment
                if len(bytes(data_col, 'GBK')) > max_len[idx][col]:
                    max_len[idx][col] = len(bytes(data_col, 'GBK'))
        for colidx in range(len(data[idx].get('titles', []))):
            cur_col = ord('A') + colidx
            ws.column_dimensions['%s' % chr(cur_col)].width = max_len[idx][colidx]
    wb.save('%s' % export_filename)

def get_day_first_timestamp(date):
    date = str(date)
    if len(date) == 10:
        return '%s 00:00:00'%date
    else:
        return date

def get_day_last_timestamp(date):
    date = str(date)
    if len(date) == 10:
        return '%s 23:59:59'%date
    else:
        return date

def gen_myktv_api_sign(bank_account, bank_name):
    sign_str = ''.join(list(map(str, [MYKTV_SECRET, bank_account, bank_name])))
    mysign = hashlib.md5(sign_str.encode()).hexdigest()
    return mysign

def check_sign(timstamp, ktv_id, sign):
    sign_str = ''.join(list(map(str, [SECURET, timstamp, ktv_id])))
    mysign = hashlib.md5(sign_str.encode()).hexdigest()
    return mysign == sign

